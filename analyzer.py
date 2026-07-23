import openpyxl
import numpy as np
import pandas as pd
from collections import defaultdict
from datetime import datetime
from typing import Optional


COLUMN_KEYWORDS = {
    "canal_dist": ["canal distribucion", "canal dist"],
    "documento": ["nro documento", "nro doc", "documento"],
    "sucursal_despacho": ["desc. sucursal despacho", "sucursal despacho"],
    "valor_subtotal": ["valor subtotal local", "valor subtotal"],
    "vendedor": ["nombre vendedor", "vendedor"],
    "co": ["c.o.", "co"],
    "valor_neto": ["valor neto"],
    "estado_docto": ["estado"],
    "valor_neto_local": ["valor neto local"],
    "cliente_factura": ["razon social cliente factura", "razon social cliente"],
    "notas_dev": ["notas causal dev", "notas causal"],
    "linea": ["linea", "línea"],
    "grupo": ["grupo"],
    "sub_linea": ["sub-linea", "sub linea", "sublínea"],
    "proveedor": ["proveedor"],
    "cliente_despacho": ["razon social cliente despacho", "cliente despacho"],
    "margen": ["margen promedio", "margen"],
    "fecha": ["fecha"],
    "sucursal_factura": ["desc. sucursal factura", "sucursal factura"],
    "pais": ["desc. pa�s", "desc. pais"],
    "desc_item": ["desc. item", "desc item", "descripcion item"],
    "ciudad": ["desc. ciudad", "ciudad"],
    "desc_co": ["desc. c.o.", "desc c.o.", "desc. co"],
    "barrio": ["desc. barrio", "barrio"],
    "costo_uni": ["costo promedio uni. inst", "costo promedio uni"],
    "costo_total": ["costo promedio total"],
    "um": ["u.m."],
    "precio_unit": ["precio unit.", "precio unitario"],
    "desc_un": ["desc. u.n.", "desc. un"],
    "tipo_cliente": ["desc. tipo de cliente", "tipo de cliente"],
    "utilidad": ["utilidad promedio", "utilidad"],
    "desc_um": ["desc. u.m.", "desc. um"],
    "tipo_docto": ["desc. tipo docto", "tipo docto"],
    "valor_descuentos": ["valor descuentos", "descuento"],
    "email": ["e-mail", "email"],
    "cantidad_inv": ["cantidad inv.", "cantidad inv"],
    "cantidad": ["cantidad"],
    "canal": ["canal"],
    "tipo_distribuidor": ["tipo distribuidor"],
    "tipo_cliente_2": ["tipo de cliente"],
    "categoria": ["categoria"],
    "estado_prod": ["estado"],
    "costo_cif": ["costo cif"],
}


def _normalize(text: str) -> str:
    result = text.strip().lower().replace(".", " ").replace("_", " ")
    while "  " in result:
        result = result.replace("  ", " ")
    return result


def _match_column(header: str, keywords: list[str]) -> bool:
    norm = _normalize(header)
    return any(_normalize(kw) in norm for kw in keywords)


def detect_columns(headers: list[str]) -> dict[str, Optional[int]]:
    detected = {}
    used_indices = set()

    priority_keys = [
        "canal_dist", "documento", "sucursal_despacho", "valor_subtotal",
        "vendedor", "co", "valor_neto", "estado_docto", "valor_neto_local",
        "cliente_factura", "linea", "grupo", "sub_linea", "proveedor",
        "cliente_despacho", "margen", "fecha", "sucursal_factura", "pais",
        "desc_item", "ciudad", "desc_co", "barrio", "costo_uni", "costo_total",
        "um", "precio_unit", "desc_un", "tipo_cliente", "utilidad", "desc_um",
        "tipo_docto", "valor_descuentos", "email", "cantidad_inv", "cantidad",
        "canal", "tipo_distribuidor", "tipo_cliente_2", "categoria", "costo_cif",
    ]

    for key in priority_keys:
        detected[key] = None
        keywords = COLUMN_KEYWORDS.get(key, [])
        for idx, h in enumerate(headers):
            if idx in used_indices:
                continue
            if h and _match_column(str(h), keywords):
                detected[key] = idx
                used_indices.add(idx)
                break

    for key in COLUMN_KEYWORDS:
        if key not in detected:
            detected[key] = None
            keywords = COLUMN_KEYWORDS[key]
            for idx, h in enumerate(headers):
                if idx in used_indices:
                    continue
                if h and _match_column(str(h), keywords):
                    detected[key] = idx
                    used_indices.add(idx)
                    break

    return detected


def read_pivot_table(ws) -> dict:
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        rows.append(list(row))

    title = None
    filter_value = None
    headers_row = None
    data_rows = []

    for i, row in enumerate(rows):
        non_none = [c for c in row if c is not None]
        if len(non_none) >= 1 and isinstance(non_none[0], str) and ("etiqueta" in _normalize(str(non_none[0])) or "fila" in _normalize(str(non_none[0]))):
            headers_row = row
            continue
        if any(c and isinstance(c, str) and "total" in _normalize(str(c)) for c in row if c):
            continue
        if headers_row is not None and any(c is not None for c in row):
            data_rows.append(row)

    return {
        "title": title,
        "filter_value": filter_value,
        "headers": headers_row,
        "data": data_rows,
    }


def _extract_pivot_filters(filepath: str) -> list[dict]:
    import zipfile
    import xml.etree.ElementTree as ET

    filters = []
    ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

    try:
        with zipfile.ZipFile(filepath) as z:
            cache_files = [n for n in z.namelist()
                          if 'pivotcachedefinition' in n.lower() and n.endswith('.xml')]
            pt_files = [n for n in z.namelist()
                       if 'pivottable' in n.lower() and 'cache' not in n.lower() and n.endswith('.xml')]

            field_names = []
            shared_items_map = {}
            for cf in cache_files:
                cf_root = ET.fromstring(z.read(cf))
                for i, cf_elem in enumerate(cf_root.findall('.//x:cacheField', ns)):
                    fname = cf_elem.get('name', f'Field_{i}')
                    field_names.append(fname)
                    si = cf_elem.find('x:sharedItems', ns)
                    if si is not None:
                        items = []
                        for child in si:
                            v = child.get('v')
                            if v is not None:
                                items.append(v)
                        shared_items_map[i] = items

            for pt in pt_files:
                pt_root = ET.fromstring(z.read(pt))
                for pfield in pt_root.findall('.//x:pivotField', ns):
                    if pfield.get('axis') != 'axisPage':
                        continue
                    items_elem = pfield.find('x:items', ns)
                    if items_elem is None:
                        continue

                    pf_elems = pt_root.findall('.//x:pageFields/x:pageField', ns)
                    fname = 'Unknown'
                    fld_idx = -1
                    for pf_elem in pf_elems:
                        fld_idx = int(pf_elem.get('fld', -1))
                        if 0 <= fld_idx < len(field_names):
                            fname = field_names[fld_idx]
                            break

                    field_shared = shared_items_map.get(fld_idx, [])

                    selected = []
                    hidden_vals = []
                    all_values = []
                    for item in items_elem.findall('x:item', ns):
                        x = item.get('x')
                        h = item.get('h')
                        t = item.get('t')
                        if t == 'default' or x is None:
                            continue
                        x_idx = int(x)
                        val = field_shared[x_idx] if x_idx < len(field_shared) else str(x_idx)
                        all_values.append(val)
                        if h == '1':
                            hidden_vals.append(val)
                        else:
                            selected.append(val)

                    if selected or hidden_vals:
                        filters.append({
                            'field_name': fname,
                            'selected': selected,
                            'hidden': hidden_vals,
                            'all_values': all_values,
                        })

    except Exception:
        return []

    return filters


def _safe_sum(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns:
        return 0.0
    return float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum())


def _safe_mean(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns:
        return 0.0
    val = pd.to_numeric(df[col], errors="coerce").fillna(0).mean()
    return float(val) if pd.notna(val) else 0.0


def _safe_sum_from_series(series: pd.Series) -> float:
    return float(pd.to_numeric(series, errors="coerce").fillna(0).sum())


def compute_dimension_metrics(df: pd.DataFrame, dim_col: str) -> list[dict]:
    if dim_col not in df.columns or df[dim_col].isna().all():
        return []

    df_valid = df.dropna(subset=[dim_col]).copy()
    for col in [dim_col]:
        if col in df_valid.columns:
            df_valid[col] = df_valid[col].astype(str).str.strip()
            df_valid[col] = df_valid[col].replace({"": None, "nan": None, "None": None})
    df_valid = df_valid.dropna(subset=[dim_col])

    results = []
    for dim_name, group in df_valid.groupby(dim_col):
        dim_name = str(dim_name).strip()
        if not dim_name or dim_name == "None":
            continue

        valor_subtotal = _safe_sum(group, "valor_subtotal")
        valor_neto = _safe_sum(group, "valor_neto")
        valor_neto_local = _safe_sum(group, "valor_neto_local")
        cantidad = _safe_sum(group, "cantidad")
        costo_total = _safe_sum(group, "costo_total")
        costo_cif = _safe_sum(group, "costo_cif")
        utilidad = _safe_sum(group, "utilidad")
        valor_descuentos = _safe_sum(group, "valor_descuentos")
        margen_promedio = _safe_mean(group, "margen")
        utilidad_promedio = _safe_mean(group, "utilidad")
        costo_uni_promedio = _safe_mean(group, "costo_uni")
        precio_unit_promedio = _safe_mean(group, "precio_unit")
        total_registros = len(group)

        documentos_unicos = group["documento"].dropna().nunique() if "documento" in group.columns else 0

        estado_dist = {}
        if "estado_docto" in group.columns:
            for estado, sub in group.groupby("estado_docto"):
                estado_str = str(estado).strip() if pd.notna(estado) else "Sin estado"
                estado_dist[estado_str] = {
                    "registros": len(sub),
                    "valor_neto": round(_safe_sum(sub, "valor_neto"), 0),
                    "cantidad": round(_safe_sum(sub, "cantidad"), 0),
                }

        top_ciudades = []
        if "ciudad" in group.columns:
            ciudad_data = group.dropna(subset=["ciudad"]).groupby("ciudad").agg(
                valor=("valor_neto", lambda x: _safe_sum_from_series(x)),
                cant=("cantidad", lambda x: _safe_sum_from_series(x)),
                registros=("ciudad", "count"),
            ).sort_values("valor", ascending=False).head(8)
            for ciudad_name, row_data in ciudad_data.iterrows():
                top_ciudades.append({
                    "ciudad": str(ciudad_name),
                    "valor_neto": round(row_data["valor"], 0),
                    "cantidad": round(row_data["cant"], 0),
                    "registros": int(row_data["registros"]),
                })

        top_items = []
        if "desc_item" in group.columns:
            item_data = group.dropna(subset=["desc_item"]).groupby("desc_item").agg(
                valor=("valor_neto", lambda x: _safe_sum_from_series(x)),
                cant=("cantidad", lambda x: _safe_sum_from_series(x)),
                registros=("desc_item", "count"),
            ).sort_values("valor", ascending=False).head(10)
            for item_name, row_data in item_data.iterrows():
                top_items.append({
                    "item": str(item_name),
                    "valor_neto": round(row_data["valor"], 0),
                    "cantidad": round(row_data["cant"], 0),
                    "registros": int(row_data["registros"]),
                })

        top_clientes = []
        for cliente_col in ["cliente_factura", "cliente_despacho"]:
            if cliente_col in group.columns:
                cliente_data = group.dropna(subset=[cliente_col]).groupby(cliente_col).agg(
                    valor=("valor_neto", lambda x: _safe_sum_from_series(x)),
                    cant=("cantidad", lambda x: _safe_sum_from_series(x)),
                    registros=(cliente_col, "count"),
                ).sort_values("valor", ascending=False).head(5)
                for cliente_name, row_data in cliente_data.iterrows():
                    top_clientes.append({
                        "cliente": str(cliente_name),
                        "valor_neto": round(row_data["valor"], 0),
                        "cantidad": round(row_data["cant"], 0),
                        "registros": int(row_data["registros"]),
                    })
                break

        margen_total = ((valor_neto - costo_total) / valor_neto * 100) if valor_neto > 0 else 0

        results.append({
            "nombre": dim_name,
            "valor_subtotal": round(valor_subtotal, 0),
            "valor_neto": round(valor_neto, 0),
            "valor_neto_local": round(valor_neto_local, 0),
            "cantidad": round(cantidad, 0),
            "costo_total": round(costo_total, 0),
            "costo_cif": round(costo_cif, 0),
            "utilidad_total": round(utilidad, 0),
            "utilidad_promedio": round(utilidad_promedio, 0),
            "valor_descuentos": round(valor_descuentos, 0),
            "margen_promedio": round(margen_promedio, 1),
            "margen_total": round(margen_total, 1),
            "costo_uni_promedio": round(costo_uni_promedio, 0),
            "precio_unit_promedio": round(precio_unit_promedio, 0),
            "total_registros": total_registros,
            "documentos_unicos": documentos_unicos,
            "estado_dist": estado_dist,
            "top_ciudades": top_ciudades,
            "top_items": top_items,
            "top_clientes": top_clientes,
        })

    results.sort(key=lambda x: x["valor_neto"], reverse=True)
    return results


def build_summary(metrics: list[dict]) -> dict:
    total_valor_neto = sum(m["valor_neto"] for m in metrics)
    total_valor_subtotal = sum(m["valor_subtotal"] for m in metrics)
    total_cantidad = sum(m["cantidad"] for m in metrics)
    total_costo = sum(m["costo_total"] for m in metrics)
    total_utilidad = sum(m["utilidad_total"] for m in metrics)
    total_descuentos = sum(m["valor_descuentos"] for m in metrics)
    total_registros = sum(m["total_registros"] for m in metrics)
    total_docs = sum(m["documentos_unicos"] for m in metrics)

    margen_general = ((total_valor_neto - total_costo) / total_valor_neto * 100) if total_valor_neto > 0 else 0
    utilidad_promedio = total_utilidad / len(metrics) if metrics else 0
    margen_promedio = sum(m["margen_promedio"] for m in metrics if m["margen_promedio"] > 0)
    margen_count = len([m for m in metrics if m["margen_promedio"] > 0])
    margen_promedio = margen_promedio / margen_count if margen_count > 0 else 0

    valor_por_registro = total_valor_neto / total_registros if total_registros > 0 else 0

    return {
        "total_dimensiones": len(metrics),
        "total_valor_neto": round(total_valor_neto, 0),
        "total_valor_subtotal": round(total_valor_subtotal, 0),
        "total_cantidad": round(total_cantidad, 0),
        "total_costo": round(total_costo, 0),
        "total_utilidad": round(total_utilidad, 0),
        "total_descuentos": round(total_descuentos, 0),
        "total_registros": total_registros,
        "total_documentos_unicos": total_docs,
        "margen_general": round(margen_general, 1),
        "margen_promedio": round(margen_promedio, 1),
        "utilidad_promedio": round(utilidad_promedio, 0),
        "valor_promedio_por_registro": round(valor_por_registro, 0),
    }


def get_filter_options(df: pd.DataFrame, col_map: dict) -> dict:
    options = {}
    filter_cols = {
        "canal_dist": "Canal Distribucion",
        "linea": "Linea",
        "grupo": "Grupo",
        "sub_linea": "Sub-Linea",
        "estado_docto": "Estado Documento",
        "desc_co": "Unidad Negocio",
        "desc_un": "Desc. U.N.",
        "tipo_cliente": "Tipo Cliente",
        "categoria": "Categoria",
        "tipo_distribuidor": "Tipo Distribuidor",
        "canal": "Canal",
        "vendedor": "Vendedor",
        "pais": "Pais",
        "ciudad": "Ciudad",
        "estado_prod": "Estado Prod.",
        "um": "Unidad Medida",
        "tipo_docto": "Tipo Documento",
    }

    for key, label in filter_cols.items():
        if key in df.columns:
            vals = sorted([str(v) for v in df[key].dropna().unique() if str(v).strip()])
            if vals:
                options[key] = {"label": label, "values": vals}

    return options


def build_vendedor_prompt(vendedor_data: dict, summary: dict, filters_applied: dict) -> str:
    filtros_text = "\n".join(f"- {k}: {v}" for k, v in filters_applied.items()) if filters_applied else "- Sin filtros aplicados"

    top_ciudades_text = ""
    if vendedor_data.get("top_ciudades"):
        top_ciudades_text = "\n".join(
            f"    - {c['ciudad']}: ${c['valor_neto']:,.0f} ({c['cantidad']:,.0f} und, {c['registros']} docs)"
            for c in vendedor_data["top_ciudades"]
        )
    else:
        top_ciudades_text = "    (Sin datos)"

    top_items_text = ""
    if vendedor_data.get("top_items"):
        top_items_text = "\n".join(
            f"    - {i['item']}: ${i['valor_neto']:,.0f} ({i['cantidad']:,.0f} und, {i['registros']} docs)"
            for i in vendedor_data["top_items"]
        )
    else:
        top_items_text = "    (Sin datos)"

    top_clientes_text = ""
    if vendedor_data.get("top_clientes"):
        top_clientes_text = "\n".join(
            f"    - {c['cliente']}: ${c['valor_neto']:,.0f} ({c['cantidad']:,.0f} und, {c['registros']} docs)"
            for c in vendedor_data["top_clientes"]
        )
    else:
        top_clientes_text = "    (Sin datos)"

    estados_text = ""
    if vendedor_data.get("estado_dist"):
        for estado, d in vendedor_data["estado_dist"].items():
            estados_text += f"    - {estado}: {d['registros']} docs, ${d['valor_neto']:,.0f}, {d['cantidad']:,.0f} und\n"

    valor_ratio = (vendedor_data['valor_neto'] / summary['total_valor_neto'] * 100) if summary['total_valor_neto'] > 0 else 0

    prompt = f"""Eres un analista experto en facturacion y ventas para una empresa de puertas, muebles y construccion. Genera un informe detallado y profesional para el siguiente vendedor/canal.

## FILTROS APLICADOS
{filtros_text}

## DATOS DEL VENDEDOR/CANAL
- Nombre: {vendedor_data['nombre']}

## METRICAS PRINCIPALES
- Valor neto total: ${vendedor_data['valor_neto']:,.0f}
- Valor subtotal local: ${vendedor_data['valor_subtotal']:,.0f}
- Cantidad total: {vendedor_data['cantidad']:,.0f} unidades
- Costo total: ${vendedor_data['costo_total']:,.0f}
- Costo CIF: ${vendedor_data['costo_cif']:,.0f}
- Utilidad total: ${vendedor_data['utilidad_total']:,.0f}
- Descuentos totales: ${vendedor_data['valor_descuentos']:,.0f}
- Total registros: {vendedor_data['total_registros']}
- Documentos unicos: {vendedor_data['documentos_unicos']}

## RENTABILIDAD
- Margen promedio: {vendedor_data['margen_promedio']:.1f}%
- Margen total: {vendedor_data['margen_total']:.1f}%
- Utilidad promedio: ${vendedor_data['utilidad_promedio']:,.0f}
- Costo promedio unitario: ${vendedor_data['costo_uni_promedio']:,.0f}
- Precio unitario promedio: ${vendedor_data['precio_unit_promedio']:,.0f}

## PARTICIPACION EN EL TOTAL
- Participacion del valor neto: {valor_ratio:.1f}%
- Valor neto total equipo: ${summary['total_valor_neto']:,.0f}
- Margen general del equipo: {summary['margen_general']:.1f}%
- Margen promedio del equipo: {summary['margen_promedio']:.1f}%

## ESTADOS DE DOCUMENTO
{estados_text}

## TOP CIUDADES
{top_ciudades_text}

## TOP ITEMS/PRODUCTOS
{top_items_text}

## TOP CLIENTES
{top_clientes_text}

---

## INSTRUCCIONES
Genera un informe profesional usando formato Markdown. IMPORTANTE: Usa tablas Markdown para datos numericos.

### 1. RESUMEN EJECUTIVO
Tabla resumen con metricas principales.

### 2. ANALISIS DE VENTAS
Tabla comparativa: Vendedor vs Equipo. Analisis de tendencias.

### 3. RENTABILIDAD
Tabla de margenes y utilidades. Analisis de rentabilidad.

### 4. PRODUCTOS Y LINEAS
Tabla de top items/productos. Analisis de portfolio.

### 5. CLIENTES Y GEOGRAFIA
Tabla de top clientes y ciudades. Analisis de mercado.

### 6. ESTADOS Y OPERACION
Tabla de estados de documento. Analisis operativo.

### 7. RECOMENDACIONES
Tabla con: #, Recomendacion, Accion, Prioridad, Impacto Esperado.

Sé directo, profesional y basado en datos numericos especificos."""

    return prompt


def build_hierarchical_pivot(df: pd.DataFrame) -> dict:
    if "canal_dist" not in df.columns or "vendedor" not in df.columns:
        return {"headers": [], "data": [], "hierarchy": {}}

    headers = [
        "Canal Distribucion", "Vendedor", "Cantidad",
        "Valor Subtotal", "Valor Neto", "Costo Total",
        "Utilidad", "Margen %", "Docs", "Registros"
    ]

    data_rows = []
    hierarchy = {}
    total_cant = 0
    total_sub = 0
    total_neto = 0
    total_costo = 0
    total_util = 0

    canal_groups = df.dropna(subset=["canal_dist"]).groupby("canal_dist", sort=False)

    for canal_name, canal_group in canal_groups:
        canal_name = str(canal_name).strip()
        if not canal_name:
            continue

        canal_cant = _safe_sum(canal_group, "cantidad")
        canal_sub = _safe_sum(canal_group, "valor_subtotal")
        canal_neto = _safe_sum(canal_group, "valor_neto")
        canal_costo = _safe_sum(canal_group, "costo_total")
        canal_util = _safe_sum(canal_group, "utilidad")
        canal_margen = ((canal_neto - canal_costo) / canal_neto * 100) if canal_neto > 0 else 0
        canal_docs = canal_group["documento"].dropna().nunique() if "documento" in canal_group.columns else 0
        canal_regs = len(canal_group)

        data_rows.append([
            canal_name, "(todos)",
            round(canal_cant, 0), round(canal_sub, 0), round(canal_neto, 0),
            round(canal_costo, 0), round(canal_util, 0), round(canal_margen, 1),
            canal_docs, canal_regs,
        ])

        total_cant += canal_cant
        total_sub += canal_sub
        total_neto += canal_neto
        total_costo += canal_costo
        total_util += canal_util

        vendedor_hier = {}
        vendedor_groups = canal_group.dropna(subset=["vendedor"]).groupby("vendedor", sort=False)

        for ven_name, ven_group in vendedor_groups:
            ven_name = str(ven_name).strip()
            if not ven_name:
                continue

            ven_cant = _safe_sum(ven_group, "cantidad")
            ven_sub = _safe_sum(ven_group, "valor_subtotal")
            ven_neto = _safe_sum(ven_group, "valor_neto")
            ven_costo = _safe_sum(ven_group, "costo_total")
            ven_util = _safe_sum(ven_group, "utilidad")
            ven_margen = ((ven_neto - ven_costo) / ven_neto * 100) if ven_neto > 0 else 0
            ven_docs = ven_group["documento"].dropna().nunique() if "documento" in ven_group.columns else 0
            ven_regs = len(ven_group)

            data_rows.append([
                "", ven_name,
                round(ven_cant, 0), round(ven_sub, 0), round(ven_neto, 0),
                round(ven_costo, 0), round(ven_util, 0), round(ven_margen, 1),
                ven_docs, ven_regs,
            ])

            documentos_hier = {}
            if "documento" in ven_group.columns:
                doc_groups = ven_group.dropna(subset=["documento"]).groupby("documento", sort=False)
                for doc_name, doc_group in doc_groups:
                    doc_name = str(doc_name).strip()
                    if not doc_name:
                        continue

                    doc_cant = _safe_sum(doc_group, "cantidad")
                    doc_sub = _safe_sum(doc_group, "valor_subtotal")
                    doc_neto = _safe_sum(doc_group, "valor_neto")
                    doc_costo = _safe_sum(doc_group, "costo_total")
                    doc_util = _safe_sum(doc_group, "utilidad")
                    doc_margen = ((doc_neto - doc_costo) / doc_neto * 100) if doc_neto > 0 else 0
                    doc_regs = len(doc_group)

                    items_list = []
                    if "desc_item" in doc_group.columns:
                        item_groups = doc_group.dropna(subset=["desc_item"]).groupby("desc_item", sort=False)
                        for item_name, item_group in item_groups:
                            item_name = str(item_name).strip()
                            if not item_name:
                                continue
                            items_list.append({
                                "item": item_name,
                                "cantidad": round(_safe_sum(item_group, "cantidad"), 0),
                                "valor_subtotal": round(_safe_sum(item_group, "valor_subtotal"), 0),
                                "valor_neto": round(_safe_sum(item_group, "valor_neto"), 0),
                                "costo_total": round(_safe_sum(item_group, "costo_total"), 0),
                                "utilidad": round(_safe_sum(item_group, "utilidad"), 0),
                                "registros": len(item_group),
                            })
                        items_list.sort(key=lambda x: x["valor_neto"], reverse=True)

                    documentos_hier[doc_name] = {
                        "cantidad": round(doc_cant, 0),
                        "valor_subtotal": round(doc_sub, 0),
                        "valor_neto": round(doc_neto, 0),
                        "costo_total": round(doc_costo, 0),
                        "utilidad": round(doc_util, 0),
                        "margen": round(doc_margen, 1),
                        "registros": doc_regs,
                        "items": items_list,
                    }

            vendedor_hier[ven_name] = {
                "documentos": documentos_hier,
                "canal": canal_name,
            }

        hierarchy[canal_name] = vendedor_hier

    total_margen = ((total_neto - total_costo) / total_neto * 100) if total_neto > 0 else 0
    data_rows.append([
        "Total General", "",
        round(total_cant, 0), round(total_sub, 0), round(total_neto, 0),
        round(total_costo, 0), round(total_util, 0), round(total_margen, 1),
        0, len(df),
    ])

    return {
        "headers": headers,
        "data": data_rows,
        "hierarchy": hierarchy,
    }


def process_excel(filepath: str, filter_overrides: dict = None) -> dict:
    try:
        xls = pd.ExcelFile(filepath, engine='openpyxl')
    except Exception as e:
        raise ValueError(f"No se pudo abrir el archivo Excel: {e}")

    sheet_names = xls.sheet_names
    if not sheet_names:
        raise ValueError("El archivo Excel no tiene hojas.")

    data_sheet_name = None
    pivot_sheet_name = None
    max_rows = 0

    sheet_dfs = {}
    for name in sheet_names:
        try:
            df_check = pd.read_excel(filepath, sheet_name=name, header=None, engine='openpyxl')
            sheet_dfs[name] = df_check
            row_count = len(df_check)
            if row_count > max_rows:
                max_rows = row_count
                data_sheet_name = name
        except Exception:
            continue

    for name, df_check in sheet_dfs.items():
        if name != data_sheet_name and 0 < len(df_check) < 120 and len(df_check.columns) <= 15:
            pivot_sheet_name = name
            break

    pivot_data = None
    pivot_filters = []
    if pivot_sheet_name:
        try:
            df_pivot = pd.read_excel(filepath, sheet_name=pivot_sheet_name, header=None, engine='openpyxl')
            pivot_data = read_pivot_table_from_df(df_pivot)
        except Exception:
            pivot_data = None
        try:
            pivot_filters = _extract_pivot_filters(filepath)
        except Exception:
            pivot_filters = []

    if not data_sheet_name:
        raise ValueError("No se encontro una hoja con datos.")

    headers = []
    try:
        df_raw = pd.read_excel(filepath, sheet_name=data_sheet_name, header=0, engine='openpyxl')
        headers = list(df_raw.columns)
        headers = [str(h) if h else '' for h in headers]
    except Exception as e:
        raise ValueError(f"Error al leer la hoja '{data_sheet_name}': {e}")

    if not headers:
        raise ValueError("La hoja de datos no tiene encabezados.")

    col_map = detect_columns(headers)

    rename_map = {}
    for key, idx in col_map.items():
        if idx is not None and idx < len(df_raw.columns):
            rename_map[df_raw.columns[idx]] = key
    df_raw = df_raw.rename(columns=rename_map)

    if "fecha" in df_raw.columns:
        df_raw["fecha"] = pd.to_datetime(df_raw["fecha"], errors="coerce")

    filter_options = get_filter_options(df_raw, col_map)

    df_filtered = df_raw.copy()

    if filter_overrides:
        for col_key, selected_val in filter_overrides.items():
            if selected_val and selected_val != "__ALL__" and col_key in df_filtered.columns:
                df_filtered = df_filtered[df_filtered[col_key].astype(str).str.strip() == selected_val]

    dimension_options = {
        "vendedor": "Vendedor",
        "linea": "Linea",
        "grupo": "Grupo",
        "sub_linea": "Sub-Linea",
        "canal_dist": "Canal Distribucion",
        "desc_co": "Unidad Negocio",
        "categoria": "Categoria",
        "canal": "Canal",
        "ciudad": "Ciudad",
        "pais": "Pais",
        "cliente_factura": "Cliente Factura",
        "cliente_despacho": "Cliente Despacho",
        "estado_docto": "Estado Documento",
        "tipo_cliente": "Tipo Cliente",
        "tipo_distribuidor": "Tipo Distribuidor",
        "desc_item": "Producto",
    }

    active_dimensions = {}
    for dim_key, dim_label in dimension_options.items():
        if dim_key in df_filtered.columns and not df_filtered[dim_key].isna().all():
            metrics = compute_dimension_metrics(df_filtered, dim_key)
            if metrics:
                active_dimensions[dim_key] = {
                    "label": dim_label,
                    "metrics": metrics,
                    "summary": build_summary(metrics),
                }

    primary_dim = "vendedor"
    if primary_dim not in active_dimensions:
        for dk in active_dimensions:
            primary_dim = dk
            break

    primary_metrics = active_dimensions.get(primary_dim, {}).get("metrics", [])
    team_summary = active_dimensions.get(primary_dim, {}).get("summary", {})

    date_range = {}
    if "fecha" in df_filtered.columns:
        fechas = df_filtered["fecha"].dropna()
        if len(fechas) > 0:
            date_range = {
                "min": str(fechas.min().date()),
                "max": str(fechas.max().date()),
            }

    pivot_summary = []
    if pivot_data and pivot_data.get("headers") and pivot_data.get("data"):
        headers_raw = pivot_data["headers"]
        col_labels = [str(h) if h else "" for h in headers_raw]
        for row in pivot_data["data"]:
            entry = {}
            for ci, h in enumerate(col_labels):
                if ci < len(row):
                    val = row[ci]
                    if pd.isna(val) if not isinstance(val, str) else False:
                        entry[h] = None
                    elif isinstance(val, str) and "(en blanco)" in val.lower():
                        entry[h] = None
                    else:
                        entry[h] = val
            label_val = entry.get(col_labels[0]) if col_labels else None
            if entry and label_val and label_val is not None and str(label_val).strip():
                pivot_summary.append(entry)

    hierarchical_pivot = build_hierarchical_pivot(df_filtered)

    return {
        "active_dimensions": active_dimensions,
        "primary_dimension": primary_dim,
        "filter_options": filter_options,
        "filter_overrides": filter_overrides or {},
        "col_map": {k: v for k, v in col_map.items() if v is not None},
        "detected_columns": {k: (headers[v] if v is not None else None) for k, v in col_map.items()},
        "total_rows": len(df_filtered),
        "total_unfiltered_rows": len(df_raw),
        "date_range": date_range,
        "pivot_summary": pivot_summary,
        "pivot_sheet": pivot_sheet_name,
        "pivot_table": hierarchical_pivot,
        "_df": df_filtered,
    }


def get_table_data(
    df: pd.DataFrame,
    page: int = 1,
    page_size: int = 50,
    sort_column: str = None,
    sort_ascending: bool = True,
    search: str = None,
) -> dict:
    display_cols = [
        "fecha", "canal_dist", "linea", "grupo", "sub_linea", "categoria",
        "canal", "desc_co", "desc_un", "tipo_cliente", "tipo_distribuidor",
        "vendedor", "ciudad", "pais", "cliente_factura", "cliente_despacho",
        "desc_item", "documento", "estado_docto", "tipo_docto",
        "cantidad", "um", "precio_unit", "costo_uni", "costo_total", "costo_cif",
        "valor_subtotal", "valor_neto", "valor_neto_local", "valor_descuentos",
        "margen", "utilidad",
    ]
    available = [c for c in display_cols if c in df.columns]
    df_show = df[available].copy()

    if search:
        search_lower = search.lower()
        mask = pd.Series([False] * len(df_show))
        for col in df_show.columns:
            try:
                col_str = df_show[col].astype(str).str.lower()
                mask = mask | col_str.str.contains(search_lower, na=False)
            except Exception:
                pass
        df_show = df_show[mask]

    total_rows = len(df_show)

    if sort_column and sort_column in df_show.columns:
        try:
            if pd.api.types.is_numeric_dtype(df_show[sort_column]):
                df_show = df_show.sort_values(sort_column, ascending=sort_ascending, na_position='last')
            else:
                df_show = df_show.sort_values(sort_column, ascending=sort_ascending, na_position='last', key=lambda x: x.astype(str).str.lower())
        except Exception:
            pass

    total_pages = max(1, (total_rows + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    end = start + page_size
    page_df = df_show.iloc[start:end]

    col_labels = {
        "fecha": "Fecha", "canal_dist": "Canal Distribucion", "linea": "Linea",
        "grupo": "Grupo", "sub_linea": "Sub-Linea", "categoria": "Categoria",
        "canal": "Canal", "desc_co": "Unidad Negocio", "desc_un": "Desc. U.N.",
        "tipo_cliente": "Tipo Cliente", "tipo_distribuidor": "Tipo Distribuidor",
        "vendedor": "Vendedor", "ciudad": "Ciudad", "pais": "Pais",
        "cliente_factura": "Cliente Factura", "cliente_despacho": "Cliente Despacho",
        "desc_item": "Producto", "documento": "Documento", "estado_docto": "Estado",
        "tipo_docto": "Tipo Docto", "cantidad": "Cantidad", "um": "U.M.",
        "precio_unit": "Precio Unit.", "costo_uni": "Costo Uni.",
        "costo_total": "Costo Total", "costo_cif": "Costo CIF",
        "valor_subtotal": "Valor Subtotal", "valor_neto": "Valor Neto",
        "valor_neto_local": "Valor Neto Local", "valor_descuentos": "Descuentos",
        "margen": "Margen", "utilidad": "Utilidad",
    }

    headers = []
    for c in available:
        headers.append({"key": c, "label": col_labels.get(c, c)})

    rows = []
    for _, row in page_df.iterrows():
        r = {}
        for c in available:
            val = row[c]
            if pd.isna(val):
                r[c] = None
            elif pd.api.types.is_datetime64_any_dtype(type(val)) or (hasattr(val, 'strftime')):
                try:
                    r[c] = val.strftime("%Y-%m-%d")
                except Exception:
                    r[c] = str(val)
            elif isinstance(val, (np.integer,)):
                r[c] = int(val)
            elif isinstance(val, (np.floating,)):
                r[c] = round(float(val), 2)
            else:
                r[c] = str(val)
        rows.append(r)

    return {
        "headers": headers,
        "rows": rows,
        "total_rows": total_rows,
        "total_pages": total_pages,
        "current_page": page,
        "page_size": page_size,
    }


def read_pivot_table_from_df(df: pd.DataFrame) -> dict:
    headers_row = None
    data_rows = []

    for i, row in df.iterrows():
        vals = [row[j] for j in range(len(df.columns))]
        non_none = [c for c in vals if pd.notna(c)]
        if any(pd.notna(c) and isinstance(c, str) and "etiqueta" in _normalize(str(c)) for c in vals):
            headers_row = [c if pd.notna(c) else None for c in vals]
            continue
        if any(pd.notna(c) and isinstance(c, str) and "total" in _normalize(str(c)) for c in vals):
            continue
        if headers_row is not None and any(pd.notna(c) for c in vals):
            data_rows.append([c if pd.notna(c) else None for c in vals])

    return {
        "title": None,
        "filter_value": None,
        "headers": headers_row,
        "data": data_rows,
    }
